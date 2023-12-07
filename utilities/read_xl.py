from openpyxl import load_workbook

# Test data is xlsx file with following format
# 
# param1        param2      param3      ...     valid       testcase name
# -----------------------------------------------------------------------
# value1        value2      value3      ...     TRUE        valid testcase
# value4        value5      value6      ...     FALSE       invalid testcase
# 
# 
# 
# FOR EXAMPLE
# 
# username      password        valid       testcase name
# --------------------------------------------------------
# student       moodle          TRUE        Valid testcase
# student       moodle123       FALSE       Wrong password
# student123    moodle          FALSE       Wrong username
#               moodle          FALSE       Empty username
# ....


# This function get all test data
def read_data_from_excel(file_name, sheet):
    datalist = []
    wb = load_workbook(filename=file_name)
    sh = wb[sheet]
    row_ct = sh.max_row
    col_ct = sh.max_column

    for i in range(2, row_ct+1):
        row = []
        for j in range(1, col_ct):
            row.append(sh.cell(row=i, column=j).value)
        datalist.append(row)
    wb.close()
    return datalist


# This function get testcase name
def get_testcase_name(file_name, sheet):
    testcase_name = []
    wb = load_workbook(filename=file_name)
    sh = wb[sheet]
    row_ct = sh.max_row
    col_ct = sh.max_column

    for i in range(2, row_ct + 1):
        cell_value = sh.cell(row=i, column=col_ct).value
        testcase_name.append(cell_value)
    wb.close()
    return testcase_name

